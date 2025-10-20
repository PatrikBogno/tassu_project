from openpyxl import load_workbook, Workbook

# Cesta k Excel súboru
subor = "predloha_2025.xlsx"

# Načítanie Excelu
wb = load_workbook(subor, data_only=True)
sheet = wb.active

# Vyberieme riadok 5 (indexovanie v Exceli začína od 1)
riadok_nazvy = 5
riadok_udaje = 7
polozka = []

bunky = [sheet.cell(row=riadok_nazvy, column=i).value for i in range(2, 176)]
bunky_cisla = [sheet.cell(row=riadok_udaje, column=i).value for i in range(2, 176)]

# Odstránime prázdne hodnoty (None alebo prázdny text)
bunky_neprazdne = [str(b).strip() for b in bunky if b not in (None, "", " ")]
kazde_druhe = bunky_cisla[1::2]
bunky_neprazdne_cisla = [str(b).strip() for b in kazde_druhe if b not in (None, "", " ")]

novy_excel = Workbook()
sheet = novy_excel.active
sheet.title = "Udaje_2025"

for i in range(len(bunky_neprazdne)):
    print(bunky_neprazdne[i] + "-" + bunky_neprazdne_cisla[i])
    sheet.cell(row=i+1, column=1, value=bunky_neprazdne[i])
    sheet.cell(row=i+1, column=2, value=bunky_neprazdne_cisla[i])

novy_excel.save("udaje_2025.xlsx")

