from openpyxl import load_workbook, Workbook

subor = "predloha_rakovina.xlsx"
wb = load_workbook(subor, data_only=True)
sheet = wb.active

stlpec_kod = 5
stlpec_pohlavie = 6
stlpec_rok = 7
stlpec_vek0_4 = 9 
stlpec_vek5_9 = 10
stlpec_vek10_14 = 11
stlpec_vek15_19 = 12
stlpec_vek20_24 = 13
stlpec_vek25_29 = 14
stlpec_vek30_34 = 15
stlpec_vek35_39 = 16
stlpec_vek40_44 = 17
stlpec_vek45_49 = 18
stlpec_vek50_54 = 19
stlpec_vek55_59 = 20
stlpec_vek60_64 = 21
stlpec_vek65_69 = 22
stlpec_vek70_74 = 23
stlpec_vek75_79 = 24
stlpec_vek80_84 = 25
stlpec_vek85_89 = 26
stlpec_vek90_94 = 27
stlpec_vek95_99 = 28

# Vytvorenie zoznamov buniek
bunky_kod = [sheet.cell(row=i, column=stlpec_kod).value for i in range(1, 625)]
bunky_pohlavie = [sheet.cell(row=i, column=stlpec_pohlavie).value for i in range(1, 625)]
bunky_rok = [sheet.cell(row=i, column=stlpec_rok).value for i in range(1, 625)]
bunky_vek0_4 = [sheet.cell(row=i, column=stlpec_vek0_4).value for i in range(1, 625)]
bunky_vek5_9 = [sheet.cell(row=i, column=stlpec_vek5_9).value for i in range(1, 625)]
bunky_vek10_14 = [sheet.cell(row=i, column=stlpec_vek10_14).value for i in range(1, 625)]
bunky_vek15_19 = [sheet.cell(row=i, column=stlpec_vek15_19).value for i in range(1, 625)]
bunky_vek20_24 = [sheet.cell(row=i, column=stlpec_vek20_24).value for i in range(1, 625)]
bunky_vek25_29 = [sheet.cell(row=i, column=stlpec_vek25_29).value for i in range(1, 625)]
bunky_vek30_34 = [sheet.cell(row=i, column=stlpec_vek30_34).value for i in range(1, 625)]
bunky_vek35_39 = [sheet.cell(row=i, column=stlpec_vek35_39).value for i in range(1, 625)]
bunky_vek40_44 = [sheet.cell(row=i, column=stlpec_vek40_44).value for i in range(1, 625)]
bunky_vek45_49 = [sheet.cell(row=i, column=stlpec_vek45_49).value for i in range(1, 625)]
bunky_vek50_54 = [sheet.cell(row=i, column=stlpec_vek50_54).value for i in range(1, 625)]
bunky_vek55_59 = [sheet.cell(row=i, column=stlpec_vek55_59).value for i in range(1, 625)]
bunky_vek60_64 = [sheet.cell(row=i, column=stlpec_vek60_64).value for i in range(1, 625)]
bunky_vek65_69 = [sheet.cell(row=i, column=stlpec_vek65_69).value for i in range(1, 625)]
bunky_vek70_74 = [sheet.cell(row=i, column=stlpec_vek70_74).value for i in range(1, 625)]
bunky_vek75_79 = [sheet.cell(row=i, column=stlpec_vek75_79).value for i in range(1, 625)]
bunky_vek80_84 = [sheet.cell(row=i, column=stlpec_vek80_84).value for i in range(1, 625)]
bunky_vek85_89 = [sheet.cell(row=i, column=stlpec_vek85_89).value for i in range(1, 625)]
bunky_vek90_94 = [sheet.cell(row=i, column=stlpec_vek90_94).value for i in range(1, 625)]
bunky_vek95_99 = [sheet.cell(row=i, column=stlpec_vek95_99).value for i in range(1, 625)]

novy_excel = Workbook()
sheet_n = novy_excel.active
sheet_n.title = "Udaje_rakovina"

k = 0

for i in range(len(bunky_kod)):
    if bunky_pohlavie[i] != "総数":
        sheet_n.cell(row=k+1, column=1, value=bunky_kod[i])
        sheet_n.cell(row=k+1, column=3, value=bunky_rok[i])
        sheet_n.cell(row=k+1, column=4, value=bunky_vek0_4[i])
        sheet_n.cell(row=k+1, column=5, value=bunky_vek5_9[i])
        sheet_n.cell(row=k+1, column=6, value=bunky_vek10_14[i])
        sheet_n.cell(row=k+1, column=7, value=bunky_vek15_19[i])
        sheet_n.cell(row=k+1, column=8, value=bunky_vek20_24[i])
        sheet_n.cell(row=k+1, column=9, value=bunky_vek25_29[i])
        sheet_n.cell(row=k+1, column=10, value=bunky_vek30_34[i])
        sheet_n.cell(row=k+1, column=11, value=bunky_vek35_39[i])
        sheet_n.cell(row=k+1, column=12, value=bunky_vek40_44[i])
        sheet_n.cell(row=k+1, column=13, value=bunky_vek45_49[i])
        sheet_n.cell(row=k+1, column=14, value=bunky_vek50_54[i])
        sheet_n.cell(row=k+1, column=15, value=bunky_vek55_59[i])
        sheet_n.cell(row=k+1, column=16, value=bunky_vek60_64[i])
        sheet_n.cell(row=k+1, column=17, value=bunky_vek65_69[i])
        sheet_n.cell(row=k+1, column=18, value=bunky_vek70_74[i])
        sheet_n.cell(row=k+1, column=19, value=bunky_vek75_79[i])
        sheet_n.cell(row=k+1, column=20, value=bunky_vek80_84[i])
        sheet_n.cell(row=k+1, column=21, value=bunky_vek85_89[i])
        sheet_n.cell(row=k+1, column=22, value=bunky_vek90_94[i])
        sheet_n.cell(row=k+1, column=23, value=bunky_vek95_99[i])

        # Pohlavie
        if bunky_pohlavie[i] == "男":
            sheet_n.cell(row=k+1, column=2, value="muz")
        else:
            sheet_n.cell(row=k+1, column=2, value="zena")

        k += 1

novy_excel.save("udaje_rakovina.xlsx")
